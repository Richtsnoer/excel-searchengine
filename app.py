from flask import Flask, jsonify, request, send_from_directory, send_file, render_template, redirect, url_for, session
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from flask_cors import CORS, cross_origin
import re  # For filename prefix stripping

# Initialize app
app = Flask(__name__)
app.secret_key = "your_secret_key_here"  # ⚡ Set a strong secret key
CORS(app, resources={r"/*": {"origins": "*"}})

# Define paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
user_file_path = os.path.join(BASE_DIR, "users.xlsx")
RDL_FOLDER = os.path.join(BASE_DIR, "rdl_files")
file_path = os.path.join(BASE_DIR, "rdlfiles.xlsx")

ALLOWED_EXTENSIONS = {"xls", "xlsx"}

# Create folders if not exist
for folder in [UPLOAD_FOLDER, RDL_FOLDER]:
    os.makedirs(folder, exist_ok=True)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def load_excel_file():
    """Load the Excel file into a DataFrame."""
    if os.path.exists(file_path):
        print(f"📂 Loading file: {file_path}")
        df = pd.read_excel(file_path, engine="openpyxl")
        df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)
        return df
    print(f"⚠️ File not found: {file_path}, creating empty DataFrame")
    return pd.DataFrame()


# Load Excel initially
df = load_excel_file()


def copy_column_widths(source_wb, target_wb):
    """Copy column widths from source workbook to target workbook."""
    source_ws = source_wb.active
    target_ws = target_wb.active
    for col in source_ws.columns:
        column_letter = get_column_letter(col[0].column)
        if column_letter in source_ws.column_dimensions:
            target_ws.column_dimensions[column_letter].width = source_ws.column_dimensions[column_letter].width


# ========================= ROUTES ==============================

from datetime import timedelta  # Add this at the top if not already

from openpyxl.utils import get_column_letter

@app.route("/upload-users", methods=["POST"])
def upload_users():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file part"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "No selected file"}), 400

        if not allowed_file(file.filename):
            return jsonify({"error": "Invalid file type. Only .xlsx files are allowed"}), 400

        # Save uploaded file temporarily
        temp_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(temp_path)
        print(f"📂 User file saved temporarily: {temp_path}")

        # Load the uploaded Excel file
        new_wb = load_workbook(temp_path)
        new_ws = new_wb.active

        # If users.xlsx exists, open it. Otherwise, create a new one
        if os.path.exists(user_file_path):
            existing_wb = load_workbook(user_file_path)
            existing_ws = existing_wb.active
        else:
            existing_wb = load_workbook(temp_path)  # Clone the first uploaded file
            existing_wb.save(user_file_path)
            new_wb.close()
            existing_wb.close()
            os.remove(temp_path)
            print("✅ Created users.xlsx from first upload.")
            return jsonify({"message": "User file created and uploaded."}), 200

        # ✅ Append new rows to the end of existing data
        new_data = list(new_ws.iter_rows(min_row=1, values_only=True))
        for row in new_data:
            existing_ws.append(row)

        # ✅ Preserve column widths
        copy_column_widths(new_wb, existing_wb)

        existing_wb.save(user_file_path)
        new_wb.close()
        existing_wb.close()
        os.remove(temp_path)

        print("✅ Users appended to users.xlsx successfully!")
        return jsonify({"message": "Users uploaded successfully!"}), 200

    except Exception as e:
        print(f"❌ Error uploading users: {e}")
        return jsonify({"error": str(e)}), 500




@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        username = request.form.get("username").strip()
        password = request.form.get("password").strip()

        try:
            if not os.path.exists(user_file_path):
                error = "User database not found."
                return render_template("login.html", error=error)

            df_users = pd.read_excel(user_file_path, engine="openpyxl").fillna("")
            df_users.columns = df_users.columns.str.strip()  # Clean column headers

            # Loop through rows to check for a match
            for _, row in df_users.iterrows():
                user = str(row.get("Username", "")).strip()
                pwd = str(row.get("Password", "")).strip()

                if user == username and pwd == password:
                    session["logged_in"] = True
                    session["username"] = username
                    return redirect(url_for("home"))

            error = "Invalid username or password"

        except Exception as e:
            error = f"Error processing login: {e}"
            print(f"❌ Login error: {e}")

    return render_template("login.html", error=error)



@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect(url_for("login"))


@app.route("/")
def home():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    username = session.get("username")
    return render_template("index.html", username=username)



@app.route("/upload", methods=["POST"])
def upload_file():
    global df

    try:
        if "file" not in request.files:
            return jsonify({"error": "No file part"}), 400

        file = request.files["file"]

        if file.filename == "":
            return jsonify({"error": "No selected file"}), 400

        if not allowed_file(file.filename):
            return jsonify({"error": "Invalid file type. Only .xlsx files are allowed"}), 400

        file_path_uploaded = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path_uploaded)
        print(f"📂 File saved to: {file_path_uploaded}")

        try:
            new_wb = load_workbook(file_path_uploaded)
            new_ws = new_wb.active
        except Exception as e:
            os.remove(file_path_uploaded)
            return jsonify({"error": f"Failed to process file: {str(e)}"}), 400

        # Load existing file
        if os.path.exists(file_path):
            existing_wb = load_workbook(file_path)
            existing_ws = existing_wb.active
        else:
            existing_wb = load_workbook(file_path_uploaded)
            existing_ws = existing_wb.active

        # Append new data
        for row in new_ws.iter_rows(values_only=True):
            existing_ws.append(row)

        copy_column_widths(new_wb, existing_wb)

        existing_wb.save(file_path)
        new_wb.close()
        existing_wb.close()

        print("✅ Upload and merge successful!")

        # Reload DataFrame
        df = load_excel_file()
        print("🔄 DataFrame reloaded after upload!")

        return jsonify({"message": "File uploaded and merged successfully!"}), 200

    except Exception as e:
        print(f"❌ ERROR in upload: {e}")
        return jsonify({"error": f"Failed to process file: {str(e)}"}), 500


@app.route("/download/<path:filename>", methods=["GET"])
def download_file(filename):
    try:
        filename = re.sub(r"^\d+\s*-\s*", "", filename)

        if not filename.lower().endswith(".rdl"):
            filename += ".rdl"

        full_path = os.path.join(RDL_FOLDER, filename)
        if not os.path.exists(full_path):
            return jsonify({"error": "File not found"}), 404

        return send_from_directory(RDL_FOLDER, filename, as_attachment=True)

    except Exception as e:
        return jsonify({"error": f"Failed to download the file: {str(e)}"}), 500


@app.route("/download-excel", methods=["GET"])
def download_excel():
    try:
        if not os.path.exists(file_path):
            return jsonify({"error": "Excel file not found"}), 404

        return send_file(file_path, as_attachment=True, download_name="rdlfiles.xlsx")

    except Exception as e:
        return jsonify({"error": f"Failed to download the file: {str(e)}"}), 500


@app.route("/search", methods=["GET"])
@cross_origin()
def search():
    global df

    try:
        query = request.args.get("query", "").strip().lower()
        if not query:
            return jsonify({"results": []})

        results = df[
            df.astype(str).apply(lambda row: row.str.contains(query, case=False, na=False).any(), axis=1)
        ]

        results_clean = results.where(pd.notnull(results), None)

        return jsonify({"results": results_clean.to_dict(orient="records")})

    except Exception as e:
        print(f"❌ Error in search route: {e}")
        return jsonify({"error": f"An error occurred while processing your request: {str(e)}"}), 500


# ================================================================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=True)
